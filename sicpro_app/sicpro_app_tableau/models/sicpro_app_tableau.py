# -*- coding: utf-8 -*-
import io

import xlsxwriter

from odoo import api, fields, models, _, http
import subprocess
from pathlib import Path

from odoo.exceptions import UserError
from odoo.http import request
from odoo.tools.safe_eval import json


class Tableau(models.Model):
    _name = 'sicpro.app.tableau'
    _description = "Vincular tablas con Tableau"

    name = fields.Char(string="Nombre", )
    trabajador = fields.Many2one('sicpro.app.trabajadores',
                                 string="Nombre del trabajador", )
    active = fields.Boolean('Activo', default=True)


    def prueba(self):
        import openpyxl

        # Tablas de modelos SICPRO
        trabajadores = self.env['sicpro.app.trabajadores'].sudo().search([('active', '=', True)])
        proveedores = self.env['sicpro.app.contratos.proveedores'].sudo().search([('active', '=', True)])

        # creo el libro excel
        wb = openpyxl.Workbook()
        hoja1 = wb.create_sheet("Trabajadores", 0)
        hoja2 = wb.create_sheet("Proveedores", 1)


        # Datos de trabajadores
        hoja1.append(('Noplaza', 'NombreApellidos', 'InicioContrato',
                      'FechaIncorporacion', 'DireccionCI', 'Raza', 'Genero',
                      'EstadoCivil', 'Hijos', 'Pasaporte', 'PermisoTrabajo',
                      'FechaSalidaPais', 'FechaRegresoPais', 'FechaBaja',
                      'NivelEscolar', 'Titulo', 'Graduacion', 'Especialidad',
                      'Donante', 'GrupoSanguineo', 'Contrato',
                      'CategoriaOcupacional', 'TelefonoTrabajo', 'MovilTrabajo',
                      'CorreoTrabajo', 'JefeInmediato', 'OcupacionLaboral',
                      'AreaTrabajo', 'CentroCosto', 'Edad',))
        celda = 1
        for item in trabajadores:
            celda = celda + 1
            hoja1.cell(row=celda, column=1, value=item.plaza_id)
            hoja1.cell(row=celda, column=2, value=item.name)
            hoja1.cell(row=celda, column=3, value=item.inicio_contrato)
            hoja1.cell(row=celda, column=4, value=item.fecha_incorporacion)
            hoja1.cell(row=celda, column=5, value=item.direccion_carnet)
            hoja1.cell(row=celda, column=6, value=item.raza)
            hoja1.cell(row=celda, column=7, value=item.genero)
            hoja1.cell(row=celda, column=8, value=item.estado_civil)
            hoja1.cell(row=celda, column=9, value=item.hijos)
            hoja1.cell(row=celda, column=10, value=item.pasaporte)
            hoja1.cell(row=celda, column=11, value=item.permiso_trabajo)
            hoja1.cell(row=celda, column=12, value=item.fecha_salida_pais)
            hoja1.cell(row=celda, column=13, value=item.fecha_regreso_pais)
            hoja1.cell(row=celda, column=14, value=item.fecha_baja)
            hoja1.cell(row=celda, column=15, value=item.nivel_escolar)
            hoja1.cell(row=celda, column=16, value=item.estudio_titulo)
            hoja1.cell(row=celda, column=17, value=item.estudio_graduacion)
            hoja1.cell(row=celda, column=18, value=item.estudio_especialidad)
            hoja1.cell(row=celda, column=19, value=item.donante)
            hoja1.cell(row=celda, column=20, value=item.grupo_sanguineo)
            hoja1.cell(row=celda, column=21, value=item.clase_contrato.name)
            hoja1.cell(row=celda, column=22, value=item.categoria_ocupacional.name)
            hoja1.cell(row=celda, column=23, value=item.telefono_trabajo)
            hoja1.cell(row=celda, column=24, value=item.movil_trabajo)
            hoja1.cell(row=celda, column=25, value=item.correo_trabajo)
            hoja1.cell(row=celda, column=26, value=item.parent_id.name)
            hoja1.cell(row=celda, column=27, value=item.ocupacion_id.name.name)
            hoja1.cell(row=celda, column=28, value=item.area_id.name)
            hoja1.cell(row=celda, column=29, value=item.centro_costo.name)
            hoja1.cell(row=celda, column=30, value=item.edad)

        # Datos de proveedores
        hoja2.append(('Nombre y Apellidos', 'Sequencia ID'))
        celda = 1
        for item in proveedores:
            celda = celda + 1

            hoja2.cell(row=celda, column=1, value=item.name)
            hoja2.cell(row=celda, column=2, value=item.sequence_consecutivo)


        wb.save('/home/daniel.borrero/SICPRO ERP/MESA DE LUZ/Odoo14/sicpro_erp/SICPRO-ERP/apps/sicpro_app_tableau/static/excel_download/BusinessIntelligenceSICPRO.xlsx')

        url = '/sicpro_app_tableau/static/excel_download/BusinessIntelligenceSICPRO.xlsx'
        return {'type': 'ir.actions.act_url', 'url': url, 'target': 'new',}


